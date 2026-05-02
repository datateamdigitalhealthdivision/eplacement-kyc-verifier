"""Build and write a manual scoring sheet for claim-guided model evaluation."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.datavalidation import DataValidation

from src.settings import AppConfig


SIGNAL_SUFFIXES = [
    "marriage",
    "self_illness",
    "family_illness",
    "spouse_location",
    "oku_self_or_family",
    "medex_other_exam",
]
MANUAL_SIGNAL_COLUMNS = [f"manual_{suffix}" for suffix in SIGNAL_SUFFIXES]
MANUAL_SIGNAL_OPTIONS = ["present", "not_present"]
MANUAL_CHECK_OPTIONS = ["check", "no_check"]


def build_scoring_sheet(decision_df: pd.DataFrame, settings: AppConfig, job_id: str) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    for _, row in decision_df.iterrows():
        scoring_row: dict[str, object] = {
            "job_id": job_id,
            "vision_model": settings.ollama.image_model,
            "secondary_vision_model": settings.ollama.secondary_image_model or "",
            "text_model": settings.ollama.model,
            "applicant_id": row.get("applicant_id", ""),
            "source_pdf_name": row.get("source_pdf_name", ""),
            "original_pdf_url": row.get("original_pdf_url", ""),
            "open_original_pdf": row.get("open_original_pdf", ""),
            "pred_check_required": row.get("check_required", ""),
            "pred_summary": row.get("summary", ""),
        }
        for suffix in SIGNAL_SUFFIXES:
            scoring_row[f"pred_{suffix}"] = row.get(f"{suffix}_status", "")
            scoring_row[f"claimed_{suffix}"] = row.get(f"claimed_{suffix}", False)
            scoring_row[f"proof_found_{suffix}"] = row.get(f"proof_found_{suffix}", False)
            scoring_row[f"verified_{suffix}"] = row.get(f"verified_{suffix}", False)
            scoring_row[f"missing_proof_{suffix}"] = row.get(f"missing_proof_{suffix}", False)
            scoring_row[f"supporting_page_{suffix}"] = row.get(f"supporting_page_{suffix}", "")
            scoring_row[f"evidence_summary_{suffix}"] = row.get(f"evidence_summary_{suffix}", "")
            scoring_row[f"confidence_{suffix}"] = row.get(f"confidence_{suffix}", 0.0)
            scoring_row[f"manual_{suffix}"] = ""
        scoring_row["manual_check_required"] = ""
        scoring_row["reviewer_notes"] = ""
        rows.append(scoring_row)
    return pd.DataFrame(rows)


def write_scoring_sheet_xlsx(scoring_df: pd.DataFrame, target_path: str | Path) -> None:
    target = Path(target_path)
    scoring_df.to_excel(target, index=False)

    workbook = load_workbook(target)
    sheet = workbook.active
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    headers = {cell.value: cell.column_letter for cell in sheet[1] if cell.value}
    last_row = sheet.max_row
    if last_row >= 2:
        manual_signal_validation = DataValidation(
            type="list",
            formula1=f'"{",".join(MANUAL_SIGNAL_OPTIONS)}"',
            allow_blank=True,
        )
        manual_check_validation = DataValidation(
            type="list",
            formula1=f'"{",".join(MANUAL_CHECK_OPTIONS)}"',
            allow_blank=True,
        )
        sheet.add_data_validation(manual_signal_validation)
        sheet.add_data_validation(manual_check_validation)
        for column_name in MANUAL_SIGNAL_COLUMNS:
            column_letter = headers.get(column_name)
            if column_letter:
                manual_signal_validation.add(f"{column_letter}2:{column_letter}{last_row}")
        check_letter = headers.get("manual_check_required")
        if check_letter:
            manual_check_validation.add(f"{check_letter}2:{check_letter}{last_row}")

        link_letter = headers.get("open_original_pdf")
        if link_letter:
            for row_number in range(2, last_row + 1):
                cell = sheet[f"{link_letter}{row_number}"]
                url = scoring_df.iloc[row_number - 2].get("open_original_pdf", "")
                if isinstance(url, str) and url.startswith(("http://", "https://")):
                    cell.value = "Open PDF"
                    cell.hyperlink = url
                    cell.font = Font(color="0563C1", underline="single")

    workbook.save(target)
